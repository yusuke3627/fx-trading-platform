"""BOE OIS spot curve collector: zip/Excel の読み取りと出所の対応付け。

配布ファイルを模した架空データを組み立てて読ませる。
"""
from __future__ import annotations

import io
import zipfile
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal

import openpyxl
import pytest

from tests.support import FakeTransport, FixedClock
from trading.data.macro.boe_yield_curve import (
    ARCHIVE_URL,
    CURRENT_MONTH_MEMBER,
    LATEST_URL,
    BOEYieldCurveCollector,
)
from trading.data.macro.registry import INDICATORS, PIT_UNVERIFIED, UK_OIS_2Y

RETRIEVED = datetime(2026, 8, 15, 3, 0, tzinfo=UTC)
YEARS = [2025, 2026]

MATURITIES = (0.5, 1, 1.5, 2, 2.5, 3)
ARCHIVE_RECENT = "OIS daily data_2025 to present.xlsx"
ARCHIVE_OLD = "OIS daily data_2016 to 2024.xlsx"


def workbook(rows: list[list], maturities: tuple = MATURITIES) -> bytes:
    """BOE の "4. spot curve" シート構成を模した workbook。"""
    book = openpyxl.Workbook()
    book.remove(book.active)
    book.create_sheet("info")
    book.create_sheet("3. spot, short end")
    sheet = book.create_sheet("4. spot curve")
    sheet.append([None, "UK OIS spot curve"])
    sheet.append([])
    sheet.append(["Maturity"])
    sheet.append(["years:", *maturities])
    # 配布ファイルには見出しの直下に計算エラーの行が入っている。
    sheet.append(["#VALUE!"])
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def zipped(members: dict[str, bytes]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        for name, data in members.items():
            archive.writestr(name, data)
    return buffer.getvalue()


def day_row(day: date, two_year: float | None) -> list:
    # 0.5 / 1 / 1.5 / 2 / 2.5 / 3 年の順。2 年は 4 列目。
    return [day, 3.9, 4.0, 4.1, two_year, 4.3, 4.4]


def collect(archive: bytes, latest: bytes, years: list[int] = YEARS, clock=None):
    transport = FakeTransport([archive, latest])
    batch = BOEYieldCurveCollector(
        transport, clock=clock or FixedClock(RETRIEVED)
    ).collect(years)
    return batch, transport


class TickingClock:
    """now() を呼ぶたび 1 分進む。取得に時間が掛かる状況の再現。"""

    def __init__(self, start: datetime) -> None:
        self._now = start

    def now(self) -> datetime:
        self._now += timedelta(minutes=1)
        return self._now


def test_takes_the_two_year_point_from_the_spot_curve() -> None:
    archive = zipped(
        {
            ARCHIVE_RECENT: workbook(
                [
                    day_row(date(2026, 7, 30), 4.18),
                    day_row(date(2026, 7, 31), 4.23),
                ]
            )
        }
    )
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([day_row(date(2026, 8, 3), 4.16)])})

    batch, transport = collect(archive, latest)

    assert [(o.observation_period, o.value) for o in batch.observations] == [
        ("2026-07-30", Decimal("4.18")),
        ("2026-07-31", Decimal("4.23")),
        ("2026-08-03", Decimal("4.16")),
    ]
    assert transport.byte_calls == [ARCHIVE_URL, LATEST_URL]


def test_the_current_month_file_wins_where_the_two_overlap() -> None:
    archive = zipped(
        {ARCHIVE_RECENT: workbook([day_row(date(2026, 8, 3), 9.99)])}
    )
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([day_row(date(2026, 8, 3), 4.16)])})

    batch, _ = collect(archive, latest)

    assert [o.value for o in batch.observations] == [Decimal("4.16")]


def test_each_day_records_the_file_it_came_from() -> None:
    archive = zipped(
        {ARCHIVE_RECENT: workbook([day_row(date(2026, 7, 31), 4.23)])}
    )
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([day_row(date(2026, 8, 3), 4.16)])})

    batch, _ = collect(archive, latest)

    by_period = {o.observation_period: o for o in batch.observations}
    assert by_period["2026-07-31"].source_uri == ARCHIVE_URL
    assert by_period["2026-08-03"].source_uri == LATEST_URL
    # payload_hash はそれぞれの配布ファイルの raw event を指す。
    hashes = {event.source_uri: event.payload_hash for event in batch.raw_events}
    assert by_period["2026-07-31"].payload_hash == hashes[ARCHIVE_URL]
    assert by_period["2026-08-03"].payload_hash == hashes[LATEST_URL]


def test_blank_values_and_non_date_rows_are_skipped() -> None:
    archive = zipped(
        {
            ARCHIVE_RECENT: workbook(
                [
                    # 休日: 日付はあるが値が空。
                    day_row(date(2026, 7, 30), None),
                    ["not a date", 1.0, 1.0, 1.0, 1.0, 1.0, 1.0],
                    day_row(date(2026, 7, 31), 4.23),
                ]
            )
        }
    )
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([])})

    batch, _ = collect(archive, latest)

    assert [o.observation_period for o in batch.observations] == ["2026-07-31"]


def test_only_the_eras_covering_the_requested_years_are_read() -> None:
    archive = zipped(
        {
            ARCHIVE_OLD: workbook([day_row(date(2024, 6, 3), 5.55)]),
            ARCHIVE_RECENT: workbook([day_row(date(2026, 7, 31), 4.23)]),
        }
    )
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([])})

    batch, _ = collect(archive, latest)

    assert [o.observation_period for o in batch.observations] == ["2026-07-31"]


def test_an_older_era_is_read_when_the_requested_years_reach_it() -> None:
    archive = zipped(
        {
            ARCHIVE_OLD: workbook([day_row(date(2024, 6, 3), 5.55)]),
            ARCHIVE_RECENT: workbook([day_row(date(2026, 7, 31), 4.23)]),
        }
    )
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([])})

    batch, _ = collect(archive, latest, years=[2024, 2025, 2026])

    assert [o.observation_period for o in batch.observations] == [
        "2024-06-03",
        "2026-07-31",
    ]


def test_rows_outside_the_requested_years_are_dropped() -> None:
    archive = zipped(
        {
            ARCHIVE_OLD: workbook(
                [
                    day_row(date(2016, 6, 3), 5.55),
                    day_row(date(2024, 6, 3), 4.44),
                ]
            ),
            ARCHIVE_RECENT: workbook([day_row(date(2026, 7, 31), 4.23)]),
        }
    )
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([])})

    batch, _ = collect(archive, latest, years=[2024, 2025, 2026])

    # member は年代単位（"_2016 to 2024" で 9 年ぶん）でしか切れないので、
    # 要求年の外の行が同じファイルに載っている。
    assert [o.observation_period for o in batch.observations] == [
        "2024-06-03",
        "2026-07-31",
    ]


def test_the_glc_workbooks_in_the_current_month_zip_are_ignored() -> None:
    archive = zipped({ARCHIVE_RECENT: workbook([])})
    latest = zipped(
        {
            "GLC Nominal daily data current month.xlsx": workbook(
                [day_row(date(2026, 8, 3), 9.99)]
            ),
            CURRENT_MONTH_MEMBER: workbook([day_row(date(2026, 8, 3), 4.16)]),
        }
    )

    batch, _ = collect(archive, latest)

    # 名目カーブを OIS として取り込むと、通貨間の減算に別物が混ざる。
    assert [o.value for o in batch.observations] == [Decimal("4.16")]


def test_a_zip_without_a_covering_workbook_raises() -> None:
    archive = zipped({ARCHIVE_OLD: workbook([])})
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([])})

    with pytest.raises(ValueError, match="no OIS workbook covering"):
        collect(archive, latest)


def test_a_curve_with_no_readable_rows_raises() -> None:
    archive = zipped({ARCHIVE_RECENT: workbook([])})
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([])})

    # 0 件を「データ無し」として通すと、欠測に気づくのは正規化が窓を
    # 満たせなくなった後になる。
    with pytest.raises(ValueError, match="yielded no observations"):
        collect(archive, latest)


def test_a_curve_without_the_two_year_column_raises() -> None:
    archive = zipped(
        {
            ARCHIVE_RECENT: workbook(
                [[date(2026, 7, 31), 3.9, 4.0, 4.1]], maturities=(0.5, 1, 1.5)
            )
        }
    )
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([])})

    with pytest.raises(ValueError, match="no 2-year column"):
        collect(archive, latest)


def test_the_series_is_pit_unverified_and_known_at_is_the_fetch_time() -> None:
    archive = zipped(
        {ARCHIVE_RECENT: workbook([day_row(date(2026, 7, 31), 4.23)])}
    )
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([])})

    batch, _ = collect(archive, latest)

    spec = INDICATORS[UK_OIS_2Y]
    assert spec.pit_classification == PIT_UNVERIFIED
    assert spec.frequency == "daily"
    assert spec.unit == "percent"
    observation = batch.observations[0]
    assert observation.known_at == RETRIEVED
    assert observation.retrieved_at == RETRIEVED
    assert observation.unit == "percent"


def test_known_at_follows_the_fetch_that_delivered_the_day() -> None:
    archive = zipped({ARCHIVE_RECENT: workbook([day_row(date(2026, 7, 31), 4.23)])})
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([day_row(date(2026, 8, 3), 4.16)])})

    batch, _ = collect(archive, latest, clock=TickingClock(RETRIEVED))

    by_period = {o.observation_period: o for o in batch.observations}
    # 取得前に時刻を打つと、まだ受け取っていない当月ぶんへ過去の known_at が
    # 付き、その間に置いた replay clock から見えてしまう。
    assert by_period["2026-07-31"].known_at < by_period["2026-08-03"].known_at
    assert by_period["2026-07-31"].known_at == RETRIEVED + timedelta(minutes=1)
    assert by_period["2026-08-03"].known_at == RETRIEVED + timedelta(minutes=2)


def test_the_workbook_itself_is_not_archived_but_its_digest_is() -> None:
    archive = zipped(
        {ARCHIVE_RECENT: workbook([day_row(date(2026, 7, 31), 4.23)])}
    )
    latest = zipped({CURRENT_MONTH_MEMBER: workbook([])})

    batch, _ = collect(archive, latest)

    payload = next(
        event.payload for event in batch.raw_events if event.source_uri == ARCHIVE_URL
    )
    assert payload["curve"] == {"2026-07-31": "4.23"}
    assert len(payload["sha256"]) == 64
